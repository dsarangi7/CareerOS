from datetime import UTC, datetime, timedelta

from openpyxl import load_workbook

from app.core.enums import ApprovalStatus
from app.models.entities import HumanApproval
from app.services.applications import (
    create_application_for_job,
    mark_applied_with_approval,
    mark_ready_to_apply,
    shortlist_job,
)
from app.services.jobs import ingest_job_description
from app.services.profile import seed_candidate_profile
from app.services.reporting import (
    export_crm_workbook,
    generate_interview_prep_pack,
    generate_weekly_report,
    overdue_followups,
    record_communication,
    record_outcome,
    schedule_follow_up,
    schedule_interview,
)

JOB_TEXT = """
Company: Phase Six Energy
Title: Senior Battery Data Engineer
Country: Canada
Required: Python, SOH analysis, BMS telemetry, and stakeholder communication.
Visa sponsorship may sponsor strong candidates.
"""


def test_phase6_reporting_workflow_reconciles_and_exports(session, tmp_path) -> None:  # type: ignore[no-untyped-def]
    seed_candidate_profile(session)
    job, *_ = ingest_job_description(session, source_text=JOB_TEXT)
    shortlist_job(session, job)
    application = create_application_for_job(session, job)
    mark_ready_to_apply(session, application)
    approval = HumanApproval(
        action_type="submit_application",
        subject_type="JobOpportunity",
        subject_id=job.id,
        status=ApprovalStatus.APPROVED,
        rationale="Phase 6 workflow test approval.",
    )
    session.add(approval)
    session.flush()
    mark_applied_with_approval(session, application, approval)

    communication = record_communication(
        session,
        application_id=application.id,
        channel="email",
        direction="inbound",
        body="Recruiter invited candidate to screening call.",
    )
    follow_up = schedule_follow_up(
        session,
        application_id=application.id,
        due_at=datetime.now(UTC) - timedelta(days=1),
        notes="Reply with availability.",
    )
    interview = schedule_interview(
        session,
        application_id=application.id,
        stage="recruiter_screen",
        scheduled_at=datetime.now(UTC) + timedelta(days=2),
    )
    pack = generate_interview_prep_pack(session, interview_id=interview.id)
    outcome = record_outcome(
        session,
        application_id=application.id,
        result="screening_scheduled",
        notes="Awaiting recruiter screen.",
    )
    report = generate_weekly_report(session, week_start="2026-08-17")
    output = export_crm_workbook(session, tmp_path / "career_os_crm.xlsx")
    session.commit()

    assert communication.id
    assert follow_up in overdue_followups(session, now=datetime.now(UTC))
    assert "company_briefing" in pack.content
    assert pack.content["questions_to_ask"]
    assert outcome.result == "screening_scheduled"
    assert report.facts["applications_total"] == 1
    assert report.facts["communications_total"] == 1
    assert report.facts["interviews_total"] == 1
    assert report.facts["followups_overdue"] == 1

    workbook = load_workbook(output)
    assert {
        "Opportunity register",
        "Application pipeline",
        "Contacts",
        "Interviews",
        "Follow-ups",
        "Outcomes",
        "Skill-gap analysis",
        "Weekly summary",
        "Data dictionary",
    }.issubset(set(workbook.sheetnames))
