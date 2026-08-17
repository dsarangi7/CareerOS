from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.enums import ApprovalStatus, JobStatus
from app.document_generation.cv import generate_tailored_cv
from app.models.entities import (
    ApplicationEvent,
    AuditEvent,
    HumanApproval,
    JobRequirement,
    ValidationResult,
)
from app.services.applications import (
    create_application_for_job,
    mark_applied_with_approval,
    mark_ready_to_apply,
    shortlist_job,
)
from app.services.jobs import ingest_job_description
from app.services.profile import seed_candidate_profile
from app.services.reporting import (
    export_crm_workbook,
    generate_interview_prep_pack,
    generate_weekly_report,
    record_communication,
    record_outcome,
    schedule_follow_up,
    schedule_interview,
)

JOB_TEXT = """
Company: Acceptance Marine Analytics
Title: Battery Diagnostics Engineer
Location: Rotterdam
Country: Netherlands
Required: Python, BMS telemetry, SOC/SOH analysis, and anomaly detection.
Required skill: stakeholder communication for battery diagnostics projects.
Experience with marine battery operational analytics.
Visa sponsorship may sponsor exceptional candidates.
"""


def test_complete_acceptance_workflow(session, tmp_path) -> None:  # type: ignore[no-untyped-def]
    profile = seed_candidate_profile(session)
    job, duplicate_of, requirements, sponsorship, fit = ingest_job_description(
        session,
        source_text=JOB_TEXT,
        source_url="https://example.test/acceptance-battery-diagnostics",
    )
    shortlist_job(session, job)
    tailored_cv = generate_tailored_cv(
        session,
        profile_id=profile.id,
        job_id=job.id,
        output_dir=tmp_path / "cv",
    )
    application = create_application_for_job(session, job)
    mark_ready_to_apply(session, application)
    approval = HumanApproval(
        action_type="submit_application",
        subject_type="JobOpportunity",
        subject_id=job.id,
        status=ApprovalStatus.APPROVED,
        rationale="Synthetic acceptance workflow approval.",
    )
    session.add(approval)
    session.flush()
    mark_applied_with_approval(session, application, approval)
    communication = record_communication(
        session,
        application_id=application.id,
        channel="email",
        direction="inbound",
        body="Recruiter response: screening invitation.",
    )
    follow_up = schedule_follow_up(
        session,
        application_id=application.id,
        due_at=datetime.now(UTC) + timedelta(days=2),
        notes="Send availability after confirming calendar.",
    )
    interview = schedule_interview(
        session,
        application_id=application.id,
        stage="screening",
        scheduled_at=datetime.now(UTC) + timedelta(days=5),
    )
    pack = generate_interview_prep_pack(session, interview_id=interview.id)
    outcome = record_outcome(
        session,
        application_id=application.id,
        result="screening_scheduled",
        notes="Synthetic outcome for acceptance test.",
    )
    report = generate_weekly_report(session, week_start="2026-08-17")
    workbook_path = export_crm_workbook(session, tmp_path / "acceptance_crm.xlsx")
    session.commit()

    assert duplicate_of is None
    assert profile.name == "Dibya Jyoti Sarangi"
    assert len(requirements) >= 2
    assert session.scalar(select(JobRequirement).where(JobRequirement.job_id == job.id)) is not None
    assert sponsorship.classification == "possibly_available"
    assert fit.total_score > 0
    assert fit.explanation
    assert job.status == JobStatus.APPLIED
    assert tailored_cv.rendered_pdf_path is not None
    assert Path(tailored_cv.rendered_pdf_path).exists()
    assert tailored_cv.validation_summary["unsupported_claims"]
    assert tailored_cv.validation_summary["finalization_blocked"] is True
    assert session.scalar(
        select(ValidationResult).where(ValidationResult.subject_id == tailored_cv.id)
    )
    assert application.status == JobStatus.APPLIED
    assert application.applied_at is not None
    assert communication.body.startswith("Recruiter response")
    assert follow_up.status == "open"
    assert pack.content["likely_screening_questions"]
    assert outcome.result == "screening_scheduled"
    assert report.facts["communications_total"] == 1
    assert report.facts["interviews_total"] == 1
    assert load_workbook(workbook_path).sheetnames
    assert session.scalar(
        select(ApplicationEvent).where(ApplicationEvent.application_id == application.id)
    )
    assert session.scalar(select(AuditEvent)) is not None
